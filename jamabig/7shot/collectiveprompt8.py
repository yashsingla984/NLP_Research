import argparse
import csv
import cloudscraper 
import requests
from bs4 import BeautifulSoup
import pandas as pd
from requests import get
import numpy as np
import time
import random
import re
from transformers import AutoTokenizer, AutoModelForCausalLM
import google.generativeai as palm
import torch
from tqdm import tqdm
import together

import openai 
import openpyxl

max_retries = 3
  
openai.api_key = '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'
token = "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"
palm.configure(api_key='AIzaSyAILWnvzeDIDPrOuzMahvRPUG7RTniujv8')

together.api_key = '74c6f051adb52d62603950542c58dc556c8a38abb68a96336e9cdd0210d46e01'


from openpyxl import Workbook


def create_excel_file(file_name):
    # Create a new Workbook object
    wb = Workbook()

    # Choose the active worksheet (sheet)
    ws = wb.active


    # Save the workbook with the provided file name
    wb.save(f'{file_name}.xlsx')
    print(f'File "{file_name}.xlsx" created successfully.')


def renamingfield(orig_name):
    if orig_name=="JAMA Cardiology Clinical Challenge ":
        #print("yashhhh")
        return "Cardiology"
    
    elif orig_name=="JAMA Cardiology Diagnostic Test Interpretation ":
        return "Cardiology"

    elif orig_name=="JAMA Clinical Challenge ":
        return "JAMA"

    elif orig_name=="JAMA Dermatology Clinicopathological Challenge ":
        return "Dermatology"

    elif orig_name=="JAMA Diagnostic Test Interpretation ":
        return "Diagnostic"

    elif orig_name=="JAMA Oncology Clinical Challenge ":
        return "Oncology"

    elif orig_name=="JAMA Oncology Diagnostic Test Interpretation ":
        return "Oncology"

    elif orig_name=="JAMA Ophthalmology Clinical Challenge ":
        return "Ophthalmology"

    elif orig_name=="JAMA Pediatrics Clinical Challenge ":
        return "Pediatrics"

    elif orig_name=="JAMA Psychiatry Clinical Challenge ":
        return "Psychiatry"
        
    elif orig_name=="JAMA Surgery Clinical Challenge ":
        return "Surgery"
    elif orig_name=="JAMA Neurology Clinical Challenge ":
        return "Neurology"

    elif orig_name=="Pediatric Quality Measures":
        #print("yassss")
        return "Pediatrics"
    
    else:
        return orig_name
    
    
    return "JAMA"

def gpt_re_answer(output):
    extracted_answer=''
    answer_pattern = re.compile(r'^([A-D])\)')
    match = answer_pattern.search(output)
    if match:
        extracted_answer = match.group(1)
        print("Extracted answer:", extracted_answer)
    if match==None:
        answer_pattern = re.compile(r'^([A-D]):')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)
    
    if match==None:
        answer_pattern = re.compile(r'Answer:\s*([A-D])')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)
    
    if match==None:
        answer_pattern = re.compile(r'\*\*Reasoning\*\*:\s(.+?)\*\*Answer\*\*: ([A-D])')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(2)
            print("Extracted answer:", extracted_answer)

    if match==None:
        answer_pattern = re.compile(r'\*\*Answer\*\*: ([A-D])')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)

    if match==None:
        answer_pattern = re.compile(r'Therefore, the correct answer is: ([A-D])')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)

    return extracted_answer


def llama_re_answer(output):
    extracted_answer=''
    answer_pattern = re.compile(r'Answer:\s*([A-D])\.\s')
    match = answer_pattern.search(output)
    if match:
        extracted_answer = match.group(1)
        print("Extracted answer:", extracted_answer)
    if match==None:
        answer_pattern = re.compile(r'Answer:\s*\(([A-D])\)')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)
    
    if match==None:
        answer_pattern = re.compile(r'Answer:\s*([A-D])')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)
    
    if match==None:
        answer_pattern = re.compile(r"\*\*Answer\*\*: \((A|B|C|D)\)")
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)

    if match==None:
        answer_pattern = re.compile(r"\*\*Answer\*\*:\((A|B|C|D)\)")
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)

#for RY
    if match==None:
        answer_pattern = re.compile(r"\*\*Answer\*\*: ([A-D])")
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)

    return extracted_answer

def palm_re_answer(output):
    extracted_answer=''
    answer_pattern = re.compile(r'Answer:\s([A-D])\.')
    match = answer_pattern.search(output)
    if match:
        print("Ist one")
        extracted_answer = match.group(1)
        print("Extracted answer:", extracted_answer)
    if match==None:
        print("2st one")
        answer_pattern = re.compile(r'Answer:\s([A-D])')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)
    
    if match==None:
        print("3st one")
        answer_pattern = re.compile(r'\(([A-D])\)') 
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)
    
    if match==None:
        print("4st one")
        answer_pattern = re.compile(r'The correct answer is ([A-D])\.')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)

    if match==None:
        print("5st one")
        answer_pattern = re.compile(r'The correct answer is: ([A-D])\)')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)

    if match==None:
        print("6st one")
        answer_pattern = re.compile(r'The answer is \*\*([A-D])')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)

    if match==None:
        print("7st one")
        answer_pattern = re.compile(r'The correct answer is: ([A-D])')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)

    if match==None:
        print("8st one")
        answer_pattern = re.compile(r'\*\*([A-D])\.\s[A-Za-z\s]+\*\*')
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)
#For RY
    if match==None:
        print("9st one")
        answer_pattern = re.compile(r"\*\*Answer:\*\* ([A-D])")
        match = answer_pattern.search(output)
        if match:
            extracted_answer = match.group(1)
            print("Extracted answer:", extracted_answer)

    return extracted_answer

def input_format(args, examples, prompt_type=None):
    inputs = []
    fewshot_examples = []
    if prompt_type == 'YR':
        for example in examples:
            title=example['Title']
            caseart=example['Case']
            mcqquestion=example['MCQ_question']
            question = "Title: "+title+". Case is: "+caseart+mcqquestion
            opa = example['Option1']
            opb = example['Option2']
            opc = example['Option3']
            opd = example['Option4']
            if 'palm' in args.model_name:
                inputs.append(
                    f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}. Please choose an answer and explain why that might be correct while the rest are incorrect. The output format is: Answer: (fill in the letter of the answer) Explanation: "
                )
            else:
                inputs.append(
                    f"Given the following clinical case, please choose an answer and explain why that might be correct while the rest are incorrect. The output format is: \nAnswer: (fill in the letter of the answer) \nExplanation: \n{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}."
                    # f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}, E: {ope}. Please choose an answer and explain why that might be correct while the rest are incorrect. The output format is: Answer: (fill in the letter of the answer) Explanation: "
                )
    elif prompt_type == 'R':
        for example in examples:
            title=example['Title']
            caseart=example['Case']
            mcqquestion=example['MCQ_question']
            question = "Title: "+title+". Case is: "+caseart+mcqquestion
            opa = example['Option1']
            opb = example['Option2']
            opc = example['Option3']
            opd = example['Option4']
            answer = example['Correct_option']
            inputs.append(
                    f"QUESTION: {question}\n"\
                    f"ANSWER CHOICES: \"A\": {opa}, \"B\": {opb}, \"C\": {opc}, \"D\": {opd}\n"\
                    f"ANSWER:{answer}\n"\
                    f"Q:\"You are a large language model that just answered the above question. Please explain why {answer} is correct answer while the rest choices are incorrect. You should explain each choice in detail.\"\n"\
                    "A:"
                )
            # inputs.append(
            #     f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}. The correct answer is {answer}. Please explain why that is correct while the rest are incorrect."
            # )
    # elif prompt_type == 'RY':
    #     for example in examples:
    #         title=example['Title']
    #         caseart=example['Case']
    #         mcqquestion=example['MCQ_question']
    #         question = "Title: "+title+". Case is: "+caseart+mcqquestion
    #         opa = example['Option1']
    #         opb = example['Option2']
    #         opc = example['Option3']
    #         opd = example['Option4']
    #         if 'palm' in args.model_name: 
    #             inputs.append(
    #                 f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}. Please think step by step and then choose an answer. The output format is: Reasoning: Answer: (fill in the letter of the answer)"
    #             )
    #         else:
    #             inputs.append(
    #                 f"Given the following clinical case, please think step by step and then choose an answer. The output format is: \n**Reasoning**: \n**Answer**: (fill in the letter of the answer). Please strictly follow the format I gave for answering. The question is:  \n{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}."
    #                 # f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}, E: {ope}. Please think step by step and then choose an answer. The output format is: Reasoning: Answer: (fill in the letter of the answer)"
    #             )
    elif prompt_type == 'RY':
        for example in examples:
            title=example['Title']
            caseart=example['Case']
            mcqquestion=example['MCQ_question']
            question = "Title: "+title+". Case is: "+caseart+mcqquestion
            opa = example['Option1']
            opb = example['Option2']
            opc = example['Option3']
            opd = example['Option4']
            if 'palm' in args.model_name: 
                inputs.append(
                    f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}. Please think step by step and then choose an answer. The output format is: Reasoning: Answer: (fill in the letter of the answer)"
                )
            else:
                inputs.append(
                    f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}. Given the above clinical case, please think step by step and then choose an answer. The output format is: \n**Reasoning**: \n**Answer**: (fill in the letter of the answer). Please strictly follow the format I gave for answering. "
                    # f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}, E: {ope}. Please think step by step and then choose an answer. The output format is: Reasoning: Answer: (fill in the letter of the answer)"
                )
    elif prompt_type == 'Y':
        if args.few_shot != 0:
            for idx, example in enumerate(examples):
                idxes = [i for i in range(len(examples))]
                idxes.remove(idx)
                select_idx = random.sample(idxes, k=args.few_shot)
                fewshot_exps = []
                for id in select_idx:
                    title=examples[id]['Title']
                    caseart=examples[id]['Case']
                    mcqquestion=examples[id]['MCQ_question']
                    question = "Title: "+title+". Case is: "+caseart+mcqquestion
                    case = f"The following are multiple choice questions (with answers) about medical knowledge. **Question:** {question} (A) {examples[id]['Option1']} (B) {examples[id]['Option2']} (C) {examples[id]['Option3']} (D) {examples[id]['Option4']}"
                    actual_pattern = re.compile(r'\b([A-D]).', re.DOTALL)
                    actual=actual_pattern.search(examples[id]['Correct_option'])
                    answer = actual.group(1)
                    fewshot_exps.append((case, answer))
                fewshot_examples.append(fewshot_exps)
                title=example['Title']
                caseart=example['Case']
                mcqquestion=example['MCQ_question']
                question = "Title: "+title+". Case is: "+caseart+mcqquestion
                opa = example['Option1']
                opb = example['Option2']
                opc = example['Option3']
                opd = example['Option4']
                
                if 'palm' in args.model_name:
                    inputs.append(
                        f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}. please choose an answer, strictly following the output format 'Answer: (fill in the letter of the answer)'"
                    )
                else:
                    inputs.append(
                        f"Given the following clinical case, please choose an answer from the given options. The output format is: \nAnswer: (fill in the letter of the answer) \n**Question:** {question} (A) {opa} (B) {opb} (C) {opc} (D) {opd}"
                    )
                
        else:
            for example in examples:
                title=example['Title']
                caseart=example['Case']
                mcqquestion=example['MCQ_question']
                question = "Title: "+title+". Case is: "+caseart+mcqquestion
                opa = example['Option1']
                opb = example['Option2']
                opc = example['Option3']
                opd = example['Option4']
                if args.option_num == 5:
                    ope = example['ope']
                    if 'palm' in args.model_name:
                        inputs.append(
                            f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}, E: {ope}. please choose an answer, strictly following the output format 'Answer: (fill in the letter of the answer)'"
                        )
                    else:
                        inputs.append(
                            f"The following are multiple choice questions (with answers) about medical knowledge. **Question:** {question} (A) {opa} (B) {opb} (C) {opc} (D) {opd} (E) {ope}"
                        )
                else:
                    if 'palm' in args.model_name:
                        inputs.append(
                            f"{question} A: {opa}, B: {opb}, C: {opc}, D: {opd}. please choose an answer, strictly following the output format 'Answer: (fill in the letter of the answer)'"
                        )
                    else:
                        inputs.append(
                            f"Given the following clinical case, please choose an answer from the given options. The output format is: \nAnswer: (fill in the letter of the answer) \n**Question:** {question} (A) {opa} (B) {opb} (C) {opc} (D) {opd}"
                        )
    
    return inputs, fewshot_examples


# prompt LLMs
def call_gpt(message, model_name, args, fewshot_exps=None):
    messages = [ {"role": "system", "content": 
            "You are a helpful assistant that answers multiple choice questions about medical knowledge."} ]
    if fewshot_exps:
        for exp in fewshot_exps:
            messages.append(
            {"role": "user", "content": exp[0]},
            )
            messages.append(
                {"role": "assistant", "content": "**Answer**:("+exp[1]+")"},
            )
    messages.append(
        {"role": "user", "content": message},
    )
    messages.append(
        {"role": "assistant", "content": "**Answer**:("},
    )
    if args.use_chat:
        response = openai.ChatCompletion.create(
            model=model_name, messages=messages
        )
        output = response.choices[0].message.content
        print(output)
    else:
        response = openai.Completion.create(
            model=model_name,
            prompt=message
        )
        output = response.choices[0]["text"]
        print(output)

    # option_pattern = re.compile(r'^([A-D])\)')
    # option = option_pattern.match(output)
    # if option:
    #     print(option)
    #     extracted_option = option.group(1)
    #     print("Extracted option:", extracted_option)
    # else:
    #     extracted_option=""
    #     print("No option found.")
    
    if args.prompt_type=="R":
        extracted_option="It is reasoning"
    else:
        extracted_option=gpt_re_answer(output)
    

    return output,extracted_option


def call_llama2(message, tokenizer, model):
    inputs = tokenizer.encode(
        message,
        return_tensors="pt"
    )

    outputs = model.generate(
        inputs,
    )
    output = tokenizer.decode(outputs[0])
    print(output)
        
    return output



def call_llama2_togetherai(message, args, fewshot_exps=None):
    messages = ""
    model_name = "togethercomputer/" + args.model_name
    if fewshot_exps:
        for exp in fewshot_exps:
            messages += f"<human>: {exp[0]}\n<bot>:"
            messages += "**Answer**:("+exp[1]+")\n"
        messages += f"<human>: {message}\n<bot>: "
    else:
        messages += f"<human>: {message}. Please choose an answer, strictly following the output format '**Answer**:(fill in the letter of the answer)'\n<bot>:"
    #print(messages)
    print("************************************")
    print(len(messages.split()))
    if len(messages.split())>1700:
        print("yash")
        tokens = messages.split()
        diff=len(messages.split())-1700
        messages=' '.join(tokens[diff:len(messages.split())])
        print(len(messages.split()))
    try:
        response = together.Complete.create(
                    prompt = messages,
                    model = model_name,
                    max_tokens = 512,
                    temperature = 0.8,
                    top_k = 60,
                    top_p = 0.6,
                    repetition_penalty = 1.1,
                    stop = ['<human>', '\n\n\n']
                )
        output = response['output']['choices'][0]['text'].strip()
        print(output)
    except Exception as e:
        
        
        print(f"Error occurred: {e}")
        print(len(messages.split()))
        print("Llama token length exception probably")
        output="Some_Issues"
        extracted_answer="Some_Issues"
        return output,extracted_answer
    
    if args.prompt_type=="R":
        extracted_answer="It is reasoning"
    else:
        extracted_answer=llama_re_answer(output)
    return output,extracted_answer



def call_palm(message, args, fewshot_exps=None):
    examples = None
    messages = []
    if fewshot_exps:
        examples = []
        for exp in fewshot_exps:
            examples.append(
                    { 
                    "input": {"content": exp[0]},
                    "output": {"content": "Answer:("+exp[1]+")"}
                }
            )
    messages.append(
        {"author": "user", "content": message},
    )
    if args.use_chat:
        #--------------------Problem: inconsistent output formats---------------------
        response=palm.chat(
                        model='models/chat-bison-001',
                        examples=examples,
                        messages=messages, 
                        temperature=0.8, 
                        context="You are a helpful assistant that answers multiple choice questions about medical knowledge."
                    )
        try:
            output = response.candidates[0]['content']
            print(output)
        except:
            output = ''
    else:
        #--------------------Problem: most outputs are none---------------------
        prompt = ""
        if fewshot_exps:
            for exp in fewshot_exps:
                prompt += exp[0]
                prompt += "("+exp[1]+")\n"
        prompt += message
        # prompt = "The following are multiple choice questions (with answers) about medical knowledge. Question: A 27-year-old man presents for an appointment to establish care. He recently was released from prison. He has felt very fatigued and has had a cough. He has lost roughly 15 pounds over the past 3 weeks. He attributes this to intravenous drug use in prison. His temperature is 99.5¬∞F (37.5¬∞C), blood pressure is 127/68 mmHg, pulse is 100/min, respirations are 18/min, and oxygen saturation is 98% on room air. QuantiFERON gold testing is positive. The patient is started on appropriate treatment. Which of the following is the most likely indication to discontinue this patient's treatment? A: Elevated liver enzymes, B: Hyperuricemia, C: Optic neuritis, D: Peripheral neuropathy, E: Red body excretions. Answer:"
        print(prompt)
        response = palm.generate_text(
                        model='models/text-bison-001',
                        prompt=prompt,
                        temperature=0.8,
                        max_output_tokens=512,
                    )
        output = response.result
        print(output)
    
    if args.prompt_type=="R":
        extracted_answer="It is reasoning"
    else:
        extracted_answer=palm_re_answer(output)
    return output,extracted_answer






















# def call_palm(input, args):
#     extracted_answer=''
#     answer_pattern = re.compile(r'Answer:\s([A-D])\.')
#     if args.use_chat:
#         response=palm.chat(
#                         model='models/chat-bison-001',
#                         messages=input, 
#                         temperature=0.8, 
#                         context="Speak like a clinician"
#                     )
#         try:
#             output = response.candidates[0]['content']
#         except:
#             output = ''

#     else:
#         response = palm.generate_text(
#                         model='models/text-bison-001',
#                         prompt=input,
#                         temperature=0.8,
#                         max_output_tokens=512,
#                     )
#         output = response.result
#     # match = answer_pattern.search(output)
#     # if match:
#     #     extracted_answer = match.group(1)
#     #     print("Extracted answer:", extracted_answer)
#     # print(output)
#     extracted_answer=palm_re_answer(output)
#     return output,extracted_answer


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--prompt_type",
        default='YR',
        type=str,
        help="YR: output answer and explanation; R: output explanation given correct answer; RY: chain-of-though; Y: answer",
    )
    parser.add_argument(
        "--model_name",
        default="gpt-3.5-turbo",
        type=str,
        help="model name: gpt-3.5-turbo, llama-2-70b-chat, gpt-3.5-turbo-instruct, palm2, gpt-4, meditron",
    )
    # parser.add_argument(
    #     "--device", default="gpu", type=str, help="GPU number or 'cpu'."
    # )
    # parser.add_argument(
    #     "--data_name",
    #     default="medbullets",
    #     type=str,
    #     help="data name: xmedqa, medbullets",
    # )
    # parser.add_argument(
    #     "--input_file",
    #     default="medbullets_test.csv",
    #     type=str,
    #     help="input file: train2_sample.csv, medbullets_test.csv",
    # )
    # parser.add_argument(
    #     "--prompt_file",
    #     default="medbullets_prompt.csv",
    #     type=str,
    #     help="prompt file",
    # )
    parser.add_argument(
        "--output_file",
        default="explanations",
        type=str,
        help="output file",
    )
    parser.add_argument(
        "--use_chat",
        default=True,
        action="store_true",
        help="Use the chat version.",
    )
    parser.add_argument(
        "--few_shot",
        default=0,
        type=int,
        help="Number of few-shot examples.",
    )
    parser.add_argument(
        "--option_num",
        default=4,
        type=int,
        help="Number of options.",
    )
    args = parser.parse_args()

    print('model_name: {}'.format(args.model_name))
    print('prompt_type: {}'.format(args.prompt_type))

    # # Setup device
    # args.device = torch.device(
    #     f"cuda"
    #     if torch.cuda.is_available() and args.device != "cpu"
    #     else "cpu"
    # )
    
    # read data
    # current_path = os.path.dirname(os.path.abspath(__file__))
    # in_file = os.path.join(current_path, 'data', args.data_name, args.input_file)
    #in_file='output_fin_jama_big2_5SHOT_repeat.xlsx'
    in_file='input_jama2.xlsx'
    df = pd.read_excel(in_file)

    examples = []
    ii = 0
    for _, row in df.iterrows():
        ii += 1
        if ii == 10:
            break
        examples.append(row)
        #print(examples)
        #print(examples[0]['Title'])
        #time.sleep(5)
    

    
    inputs, fewshot_examples = input_format(args, examples, args.prompt_type)

    if 'llama-2' in args.model_name:
        # # load model from huggingface
        # # tokenizer = AutoTokenizer.from_pretrained(
        # #     "meta-llama/"+args.model_name
        # # )
        # # model = AutoModelForCausalLM.from_pretrained(
        # #     "meta-llama/"+args.model_name,
        # #     torch_dtype=torch.bfloat16,
        # #     low_cpu_mem_usage=True,
        # # )

        # # outputs = []
        # # for input in tqdm(inputs):
        # #     output = call_llama2(input, tokenizer, model)
        # #     outputs.append(output)

        # # together ai
        # model_name = "togethercomputer/" + args.model_name
        # outputs = []
        # extracted_options=[]
        # for input in tqdm(inputs):
        #     response = together.Complete.create(
        #         prompt = f"<human>: {input}\n<bot>:",
        #         model = model_name,
        #         max_tokens = 512,
        #         temperature = 0.8,
        #         top_k = 60,
        #         top_p = 0.6,
        #         repetition_penalty = 1.1,
        #         stop = ['<human>', '\n\n\n']
        #     )
        #     print(response['output'])
        #     output = response['output']['choices'][0]['text'].strip()
        #     print(output)
        #     extracted_option=llama_re_answer(output)
        outputs = []
        extracted_options=[]
        if len(fewshot_examples) != 0:
            for input, fewshot_exps in tqdm(zip(inputs, fewshot_examples)):
                output,extracted_option = call_llama2_togetherai(input, args, fewshot_exps)
                outputs.append(output)
                extracted_options.append(extracted_option)
        else:
            for input in tqdm(inputs):
                output,extracted_option = call_llama2_togetherai(input, args)
                outputs.append(output)       
                extracted_options.append(extracted_option)

            #outputs.append(output)

    elif 'gpt' in args.model_name:  
        # call model
        outputs = []
        extracted_options=[]
        print("yash")
        if len(fewshot_examples) != 0:
            for input, fewshot_exps in tqdm(zip(inputs, fewshot_examples)):
                output,extracted_option = call_gpt(input, args.model_name, args, fewshot_exps)
                outputs.append(output)
                extracted_options.append(extracted_option)
        else:
            for input in tqdm(inputs):
                output,extracted_option = call_gpt(input, args.model_name, args)
                outputs.append(output)
                extracted_options.append(extracted_option)
    
    # elif 'palm' in args.model_name:
    #     outputs = []
    #     extracted_options=[]
    #     for input in tqdm(inputs):
    #         try:
    #             output,extracted_option= call_palm(input, args)
    #             outputs.append(output)
    #             extracted_options.append(extracted_option)
    #             print(extracted_option)
    #         except Exception as e:
    #             output="Some_Issues"
    #             extracted_option="Some_Issues"
    #             outputs.append(output)
    #             extracted_options.append(extracted_option)


    elif 'palm' in args.model_name:
        outputs = []
        extracted_options=[]
        if len(fewshot_examples) != 0:
            for input, fewshot_exps in tqdm(zip(inputs, fewshot_examples)):
                try:
                    output,extracted_option = call_palm(input, args, fewshot_exps)
                    outputs.append(output)
                    extracted_options.append(extracted_option)
                except Exception as e:
                    output="Some_Issues"
                    extracted_option="Some_Issues"
                    outputs.append(output)
                    extracted_options.append(extracted_option)

                    
        else:
            for input in tqdm(inputs):
                try:
                    output,extracted_option= call_palm(input, args)
                    outputs.append(output)
                    extracted_options.append(extracted_option)
                    print(extracted_option)
                except Exception as e:
                    output="Some_Issues"
                    extracted_option="Some_Issues"
                    outputs.append(output)
                    extracted_options.append(extracted_option)
                # output = call_palm(input, args)
                # outputs.append(output)



    # elif 'meditron' in args.model_name:
    #     # load model from huggingface
    #     tokenizer = AutoTokenizer.from_pretrained(
    #         "epfl-llm/meditron-70b"
    #     )
    #     model = AutoModelForCausalLM.from_pretrained(
    #         "epfl-llm/meditron-70b",
    #         torch_dtype=torch.bfloat16,
    #         low_cpu_mem_usage=True,
    #     )
    #     model.to(args.device)

    #     outputs = []
    #     for input in tqdm(inputs):
    #         output = call_meditron(input, tokenizer, model, args)
    #         outputs.append(output)

    # write into file
    # filename = args.input_file.split('.')[0]
    # output_file = os.path.join(current_path, args.output_file, args.data_name+'_'+args.model_name+'_'+filename+'_'+args.prompt_type+'.csv')
    
    
    combined_name=f"{args.model_name}_{args.prompt_type}"
    
    #output_file='palm2_Y.xlsx'
    # Create a new workbook and select the active sheet
    # workbook = openpyxl.Workbook()
    # sheet = workbook.active
    if args.few_shot != 0:
        # output_file = os.path.join(current_path, args.output_file, args.data_name+'_'+args.model_name+'_'+filename+'_'+args.prompt_type+'_fs_'+\
        #                        str(args.few_shot)+'.csv')
        combined_name=f"{args.model_name}_{args.prompt_type}_{args.few_shot}"
        #output_file='yash30.xlsx'
    
    create_excel_file(combined_name)
    output_file=combined_name+".xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    if args.option_num == 5:
        header = ['question', 'opa', 'opb', 'opc', 'opd', 'opa_percent', 'opb_percent', 'opc_percent', 'opd_percent', \
                'ope_percent', 'total', 'answer_idx', 'answer', 'complexity', 'importance', 'explanation', 'input', 'output']
        combs = []
        for e, i, o in zip(examples[:len(inputs)], inputs, outputs):
            combs.append(
                [
                    e['question'],
                    e['opa'],
                    e['opb'],
                    e['opc'],
                    e['opd'],
                    e['ope'],
                    e['opa_percent'],
                    e['opb_percent'],
                    e['opc_percent'],
                    e['opd_percent'],
                    e['ope_percent'],
                    e['total'],
                    e['answer_idx'],
                    e['answer'],
                    e['complexity'],
                    e['importance'],
                    e['explanation'],
                    i,
                    o
                ]
            )
    else:
        header = ['Title','Case','MCQ_question','URL','Correct_option','answer_index', 'opa', 'opb', 'opc', 'opd','input', 'output','LLM_answer','IsLLM_Correct','MedicalField']
        sheet.append(header)

        combs = []
        for e, i, o,eo in zip(examples[:len(inputs)], inputs, outputs,extracted_options):
            actual_pattern = re.compile(r'\b([A-D]).', re.DOTALL)
            actual=actual_pattern.search(e['Correct_option'])
            isllmcorrect="No"
            try:
                if o=="Some_Issues" or eo=="Some_Issues":
                    modifiedmedicalfield=e['MedicalField']
                    cellvalue=e['Superclass']
                    if pd.isna(cellvalue):
                        modifiedmedicalfield=e['MedicalField']
                    else:
                        modifiedmedicalfield=e['Superclass']
                    
                    finalmedicalfield=renamingfield(modifiedmedicalfield)
                    row_data=[
                        e['Title'],
                        e['Case'],
                        e['MCQ_question'],
                        e['URL'],
                        e['Correct_option'],
                        "Some Issues",
                        e['Option1'],
                        e['Option2'],
                        e['Option3'],
                        e['Option4'],
                        i,
                        "Some Issues",
                        "Some Issues",
                        "No",
                        finalmedicalfield
                    ]
                    sheet.append(row_data)
                    raise Exception("An error occured")
                
                elif args.prompt_type=="R":
                    isllmcorrect="It is explanation so can't comment"
                    modifiedmedicalfield=e['MedicalField']
                    cellvalue=e['Superclass']
                    if pd.isna(cellvalue):
                        modifiedmedicalfield=e['MedicalField']
                    else:
                        modifiedmedicalfield=e['Superclass']
                    
                    finalmedicalfield=renamingfield(modifiedmedicalfield)
                    row_data=[
                            e['Title'],
                            e['Case'],
                            e['MCQ_question'],
                            e['URL'],
                            e['Correct_option'],
                            actual.group(1),
                            e['Option1'],
                            e['Option2'],
                            e['Option3'],
                            e['Option4'],
                            i,
                            o,
                            eo,
                            isllmcorrect,
                            finalmedicalfield
                        ]
                    sheet.append(row_data)


                
                else:
                    if actual.group(1)==eo:
                        isllmcorrect="Yes"
                    modifiedmedicalfield=e['MedicalField']
                    cellvalue=e['Superclass']
                    if pd.isna(cellvalue):
                        modifiedmedicalfield=e['MedicalField']
                    else:
                        modifiedmedicalfield=e['Superclass']
                    
                    finalmedicalfield=renamingfield(modifiedmedicalfield)
                    row_data=[
                            e['Title'],
                            e['Case'],
                            e['MCQ_question'],
                            e['URL'],
                            e['Correct_option'],
                            actual.group(1),
                            e['Option1'],
                            e['Option2'],
                            e['Option3'],
                            e['Option4'],
                            i,
                            o,
                            eo,
                            isllmcorrect,
                            finalmedicalfield
                        ]
                    sheet.append(row_data)
            except Exception as e:
                # row_data=[
                #         e['Title'],
                #         e['Case'],
                #         e['MCQ_question'],
                #         e['URL'],
                #         e['Correct_option'],
                #         "Some Issues",
                #         e['Option1'],
                #         e['Option2'],
                #         e['Option3'],
                #         e['Option4'],
                #         i,
                #         "Some Issues",
                #         "Some Issues",
                #         "No",
                #         finalmedicalfield
                #     ]
                # sheet.append(row_data)
                print("some exception occured")
    
    workbook.save(output_file)
    # with open(output_file, 'w', encoding='UTF8', newline='') as f:
    #     writer = csv.writer(f)
    #     writer.writerow(header)
    #     writer.writerows(combs)


if __name__ == "__main__":
    main()